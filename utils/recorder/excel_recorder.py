# -*- coding: utf-8 -*-

import contextlib
import os
import re

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet





def create_xlsx(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        print("We have created a new excel file!!!")
        Workbook().save(xlsx_path)
    else:
        print("Excel file has existed!")


@contextlib.contextmanager
def open_excel(xlsx_path: str, sheet_name: str):
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name, index=0)
    sheet = wb[sheet_name]

    yield sheet

    wb.save(xlsx_path)


def append_row(sheet: Worksheet, row_data):
    assert isinstance(row_data, (tuple, list))
    sheet.append(row_data)


def insert_row(sheet: Worksheet, row_data, row_id, min_col=1, interval=0):
   
    assert isinstance(row_id, int) and isinstance(min_col, int) and row_id > 0 and min_col > 0
    assert isinstance(row_data, (tuple, list)), row_data

    num_elements = len(row_data)
    row_data = iter(row_data)
    for row in sheet.iter_rows(
            min_row=row_id, max_row=row_id, min_col=min_col, max_col=min_col + (interval + 1) * (num_elements - 1)
    ):
        for i, cell in enumerate(row):
            if i % (interval + 1) == 0:
                sheet.cell(row=row_id, column=cell.column, value=next(row_data))


def insert_cell(sheet: Worksheet, row_id, col_id, value):
    assert isinstance(row_id, int) and isinstance(col_id, int) and row_id > 0 and col_id > 0

    sheet.cell(row=row_id, column=col_id, value=value)


def merge_region(sheet: Worksheet, min_row, max_row, min_col, max_col):
    assert max_row >= min_row > 0 and max_col >= min_col > 0

    merged_region = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    sheet.merge_cells(merged_region)


def get_col_id_with_row_id(sheet: Worksheet, col_name: str, row_id):
   
    assert isinstance(row_id, int) and row_id > 0

    for cell in sheet[row_id]:
        if cell.value == col_name:
            return cell.column
    raise ValueError(f"In row {row_id}, there is not the column {col_name}!")


def get_row_id_with_col_name(sheet: Worksheet, row_name: str, col_name: str):
    
    is_new_row = True
    col_id = get_col_id_with_row_id(sheet=sheet, col_name=col_name, row_id=1)

    row_id = 0
    for cell in sheet[get_column_letter(col_id)]:
        row_id = cell.row
        if cell.value == row_name:
            return (row_id, col_id), not is_new_row
    return (row_id + 1, col_id), is_new_row


def get_row_id_with_col_id(sheet: Worksheet, row_name: str, col_id: int):
    
    assert isinstance(col_id, int) and col_id > 0

    is_new_row = True
    row_id = 0
    for cell in sheet[get_column_letter(col_id)]:
        row_id = cell.row
        if cell.value == row_name:
            return row_id, not is_new_row
    return row_id + 1, is_new_row


def format_string_with_config(string: str, repalce_config: dict = None):
    assert repalce_config is not None

    if repalce_config.get("lower"):
        string = string.lower()
    elif repalce_config.get("upper"):
        string = string.upper()
    elif repalce_config.get("title"):
        string = string.title()

    if sub_rule := repalce_config.get("replace"):
        string = re.sub(pattern=sub_rule[0], repl=sub_rule[1], string=string)
    return string


class MetricExcelRecorder(object):
    def __init__(
            self,
            xlsx_path: str,
            sheet_name="results",
            row_header="methods",
            repalce_config=None,
            dataset_names=None,
            metric_names=None,
    ):
       
        create_xlsx(xlsx_path=xlsx_path)

        if repalce_config is None:
            repalce_config = dict(lower=True, replace=(r"[_-]", ""))
        if dataset_names is None:
            dataset_names = ["pascals", "ecssd", "hkuis", "dutste", "dutomron"]
        if metric_names is None:
            metric_names = ["smeasure", "wfmeasure", "mae", "adpfm", "meanfm", "maxfm", "adpem", "meanem", "maxem"]

        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name
        self.repalce_config = repalce_config

        self.row_header = format_string_with_config(row_header, self.repalce_config)

        self.dataset_names = [format_string_with_config(s, self.repalce_config) for s in dataset_names]
        self.metric_names = [format_string_with_config(s, self.repalce_config) for s in metric_names]
        self.num_datasets = len(self.dataset_names)
        self.num_metrics = len(self.metric_names)

        self._initial_table()

    def _initial_table(self):
        """
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |methods|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1... |
        |       |metric1      |metric2        |...|metric1      |metric2        |...|metric1      |metric2...         |
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |...
        """
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
            
            insert_cell(sheet=sheet, row_id=1, col_id=1, value=self.row_header)
       
            merge_region(sheet=sheet, min_row=1, max_row=2, min_col=1, max_col=1)
            
            insert_row(sheet=sheet, row_data=self.dataset_names, row_id=1, min_col=2, interval=self.num_metrics - 1)
          
            for i in range(self.num_datasets):
                insert_row(sheet=sheet, row_data=self.metric_names, row_id=2, min_col=2 + i * self.num_metrics)

    def _format_row_data(self, row_data: dict) -> list:
        row_data = {format_string_with_config(k, self.repalce_config): v for k, v in row_data.items()}
        return [row_data[n] for n in self.metric_names]

    def __call__(self, row_data: dict, dataset_name: str, method_name: str):
        dataset_name = format_string_with_config(dataset_name, self.repalce_config)
        assert dataset_name in self.dataset_names, f"{dataset_name} is not contained in {self.dataset_names}"

        
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
          
            dataset_col_start_id = get_col_id_with_row_id(sheet=sheet, col_name=dataset_name, row_id=1)
            (method_row_id, method_col_id), is_new_row = get_row_id_with_col_name(
                sheet=sheet, row_name=method_name, col_name="methods"
            )
           
            if is_new_row:
                sheet.cell(row=method_row_id, column=method_col_id, value=method_name)
           
            row_data = self._format_row_data(row_data=row_data)
            insert_row(sheet=sheet, row_data=row_data, row_id=method_row_id, min_col=dataset_col_start_id)


class NewMetricExcelRecorder(object):
    def __init__(
            self,
            xlsx_path: str,
            repalce_config: dict = None,
            sheet_name: str = "results",
            row_header: str = "methods",
            dataset_names: tuple = ("pascals", "ecssd", "hkuis", "dutste", "dutomron"),
            metric_names: tuple = (
                    "smeasure", "wfmeasure", "mae", "adpfm", "meanfm", "maxfm", "adpem", "meanem", "maxem"),
            dataset_lengths: tuple = (850, 1000, 4447, 5017, 5168),
            record_average: bool = True,
    ):
        assert all([isinstance(x, int) for x in dataset_lengths])
        assert len(dataset_names) == len(dataset_lengths)

        create_xlsx(xlsx_path=xlsx_path)
        self.xlsx_path = xlsx_path

        if repalce_config is None:
            self.repalce_config = dict(lower=True, replace=(r"[_-]", ""))
        else:
            self.repalce_config = repalce_config

        self.row_header = format_string_with_config(row_header, self.repalce_config)
        self.dataset_names = [format_string_with_config(s, self.repalce_config) for s in dataset_names]
        self.metric_names = [format_string_with_config(s, self.repalce_config) for s in metric_names]
        self.dataset_lengths = [float(s) for s in self.dataset_lengths]
        self.record_average = record_average

        self.num_datasets = len(self.dataset_names)
        self.num_metrics = len(self.metric_names)

        self.sheet_name = sheet_name
        self._initial_table()

    def _initial_table(self):
        """
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |methods|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1... |
        |       |metric1      |metric2        |...|metric1      |metric2        |...|metric1      |metric2...         |
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |...
        """
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
          
            insert_cell(sheet=sheet, row_id=1, col_id=1, value=self.row_header)
       
            merge_region(sheet=sheet, min_row=1, max_row=3, min_col=1, max_col=1)

            if self.record_average:
     
                self.dataset_names.append("average")
                self.dataset_lengths.append(sum(self.dataset_lengths))
                self.num_datasets += 1

      
            insert_row(sheet=sheet, row_data=self.dataset_names, row_id=1, min_col=2, interval=self.num_metrics - 1)
            insert_row(sheet=sheet, row_data=self.dataset_lengths, row_id=1, min_col=3, interval=self.num_metrics - 1)
          
            for i in range(len(self.dataset_names)):
                insert_row(sheet=sheet, row_data=self.metric_names, row_id=2, min_col=2 + i * self.num_metrics)

    def _format_row_data(self, row_data: dict) -> list:
        row_data = {format_string_with_config(k, self.repalce_config): v for k, v in row_data.items()}
        return [row_data[n] for n in self.metric_names]

    def __call__(self, row_data: dict, dataset_name: str, method_name: str):
        assert dataset_name in self.dataset_names, f"{dataset_name} is not contained in {self.dataset_names}"

        dataset_name = format_string_with_config(dataset_name, self.repalce_config)


        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
       
            dataset_col_start_id = get_col_id_with_row_id(sheet=sheet, col_name=dataset_name, row_id=1)
            (method_row_id, method_col_id), is_new_row = get_row_id_with_col_name(
                sheet=sheet, row_name=method_name, col_name=self.row_header
            )
         
            if is_new_row:
                insert_cell(sheet=sheet, row_id=method_row_id, col_id=method_col_id, value=method_name)
          
            row_data = self._format_row_data(row_data=row_data)
            insert_row(sheet=sheet, row_data=row_data, row_id=method_row_id, min_col=dataset_col_start_id)
